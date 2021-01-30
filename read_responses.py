from pathlib import Path as p
import openpyxl
from openpyxl import utils

"""
This script opens the named excel Document.
It reads and stores its values in a list 
How it works
Each row in the document is a person, and their data. For each row in the document 
    the program gets all their data recording each "Key-Value" in a dictionary. Each dictionary is
    then stored in list. As such each person is an index in a list with their info in the dictionary.
"""

loc_excel_docs = p.cwd().joinpath('ExcelDocs')  # Create a path to directory with documents

wb = openpyxl.load_workbook(p.joinpath(loc_excel_docs, 'responses.xlsx'))  # Open excel

rspns_sheet = wb['Form Responses 1']
rspns = []  # list with spread sheet response each row in a dictionary

dic_temp = {}

max_cols = rspns_sheet.max_column  # get maximum number of columns
max_rows = rspns_sheet.max_row  # get maximum number of rows

for row in range(2, max_rows):
    for i in range(2, max_cols):
        var = utils.get_column_letter(i)  # convert column num to letter

        # info to get the keys for temp dictionary [Tile used in Spread Sheet ]
        title = (rspns_sheet[(var + '1')].value[2:].strip(
            '.').strip())  # title for column( removes '.' white space and first 2 letters
        # dbs_key = None
        if '.' in title[:2]:
            """"
            if '.' in first two items of returned string return variable without the the '.' and use strip()
            """
            dbs_key = title[2:].strip()
        else:
            dbs_key = title

        # info to get values
        value = rspns_sheet[(var + str(row))].value

        dic_temp[dbs_key] = value  # add title & value to dictionary

    rspns.append(dic_temp.copy())
    dic_temp.clear()  # clear dictionary to prevent duplicate data

def update_key_list():
    # updates the text list with titles of the columns form the excel document 
    with open("Responses.txt",'r') as r:
        for key, value in rspns[1].items():
            r.write(key)
            r.write('\n')
            
            
def other ():        
    for key , value in rspns[3].items():
        print(key,"--------> ", value)

#update_key_list