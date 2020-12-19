from pathlib import Path as p
import openpyxl
from openpyxl import utils

loc_excel_docs = p.cwd().joinpath('ExcelDocs')  # Create a path to directory with documents

wb = openpyxl.load_workbook(p.joinpath(loc_excel_docs,'Database Registration.xlsx'))  # Open excel

# Below is a dict with key = Catholic Parish, value = Mass center in Parish
ParishAndMassCenter = {
   'Immac. Conc. Cathedral':
       ['Cathedral','Cathedral of the Immaculate Conception','Our Lady Queen of Peace','Our Lady Queen of the Universe',
        'Our Lady, Lily of the Valley, Tempe','Uganda Martyrs, Happy Hill',],

    'Our Lady HOC Beaulieu':
        ['Church of the Living Water, La Mode','Church of the Nativity',
         'Church of the Nativity','Our Lady Queen of Peace, Willis'],

    'Blessed Sacrament Parish':
        ['Our Lady of Mt. Carmel','Blessed Sacrament Parish'],

    'Sts JosMich Morne Jaloux':
        ['St. Joseph\'s RC','St. Michael the Archangel'],

    'Our Lady of the Rosary Roxbor':
        [],

    'St Dominic\'s':
        ['Church of the Good Shepherd','St. Dominic\'s '],

    'Immac. Conc St Joseph':
        ['church of the Immaculate Conception and St.Joseph',],
    ''
    'St MartindePorres Crochu':
        ['St MartindePorres Crochu','St Martinde Porres','St MartindePorres',
         'Crochu'],

    'Holy Cross Munich':
        ['Holy Cross RC','Holy Cross','Munich','Holy Cross RC church',
         'Our Lady of Fatima, Battle Hill','Battle Hill Shrine'],

    'St Andrew the Apostle Grenville':
        ['St Andrew the Apostle','Grenville'],

    'St Matthew Birchgrove':
        ['St Matthew Birchgrove','Birchgrove',
         'St. Matthew\'s RC','St. Matthew\'s RC Church','St. Matthew Roman Catholic'],

    'Sacred Heart Tivoli':
        ['Sacred Heart','Tivoli','Scared Heart RC','St. Gerard Moyah'],

    'Holy Family River Sallee':
        ['Holy Family River Sallee'],

    'St. Patrick\'s':
        ['Church of the Holy Spirit','St. Patrick\'s RC'],

    'St Mark the Evangelist':['St Mark the Evangelist'],

    'St Peter Gouyave':['St Peter Gouyave'],

    'Christ the King Grand Roy':['Christ the King Grand Roy'],

    'St Patrick Hillsborough':
        ['St. Margaret Mary','Holy Rosary'],

    'Sts PeterPaul Windward':['Sts PeterPaul Windward'],

    'Sacred Heart Petite Martinique':['Sacred Heart Petite Martinique']
}


catholic_parishes = []  # list of keys in ParishAndMassCenter
#  Store Parishes in list
for key, value in ParishAndMassCenter.items():
    catholic_parishes.append(str(key))

rgstn_sheet = wb[catholic_parishes[1]]

temp = []
for col in range(2,(rgstn_sheet.max_column)):
    var = utils.get_column_letter(col)

    final_value = None

    value = rgstn_sheet[(var + '1')].value # get value in top cell
    #temp.append(value)
    sub_value = rgstn_sheet[(var + '2')].value  # get value in seconday cell
    if sub_value == None:
        # sub value is none
        final_value = value
    else:
        if value == None:
            #sub vale and core valuie is NONE
            #print(temp,"TEMPP")
            value_2 = temp[-0]
            final_value = f' {value_2} : { sub_value}'
        else:
            # value and sub value have a value
            temp.clear()
            temp.append(value)
            final_value = f'{value} : {sub_value}'

    print(var + "1", str(final_value.strip()))
    with open("database.txt",'a') as doc:
        doc.write(str(final_value.strip()))
        doc.write('\n')



print(temp)



